# downloadxmldata.py
# Description: Obtains tumor stage information for a dataset by querying the clinical XML files for each patient file 
# and saving the extracted information to a new sheet in "[dataset name]-patient-data.xlsx." If no information found, 
# the error code is "no data."
# Amanda Zong, Summer 2018

import requests as rq
import json
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import pandas as pd
import sys

def main():
    # file_id = '2f1a4b16-104c-40eb-8114-8245847bb716'
    # p = 'TCGA-LUAD'
    # dataset = p[-4:].lower()
    # get_tumor_stage(file_id, dataset, 'no data')

    p = sys.argv[1]
    CASE_ID_COL = 1
    SUBMITTER_ID_COL = 2
    DEFAULT = 'no data'

    filename = f'{p}-patient-data.xlsx'
    sheetname = 'tumor_stage'
    dataset = p[5:].lower()

    wb = load_workbook(filename)
    sheet = wb.worksheets[0]

    header = ['case_id', 'submitter_id', sheetname]
    
    tumor_stage = []

    for row in range(2, sheet.max_row+1): #skips header row
        case_id = sheet.cell(row,CASE_ID_COL).value
        print(case_id)
        submitter_id = sheet.cell(row, SUBMITTER_ID_COL).value
        tumor_stage.append([case_id, submitter_id, get_tumor_stage(case_id, dataset, DEFAULT)])

    df = pd.DataFrame(data=tumor_stage)
    print("Finished compiling data.")

    writer = pd.ExcelWriter(filename, engine = 'openpyxl')
    writer.book = wb
    df.to_excel(writer, header = header, sheet_name = sheetname, index = False)
    writer.save()

def get_tumor_stage(case_id, dataset, DEFAULT):

    try:
        filters = _FilterBuilder.logical('and', [
            _FilterBuilder.equal('files.data_category', 'Clinical'),
            _FilterBuilder.equal('cases.case_id', case_id)])
        
        file_name = get_filename(filters, DEFAULT)
        #print(file_name)

        if (file_name == DEFAULT):
            return DEFAULT

        data_endpt = 'https://api.gdc.cancer.gov/data/{}'.format(file_name)
        #data_endpt = "https://api.gdc.cancer.gov/data/b8cfb22d-b445-4961-993d-ebe5ac28874a"
        response = rq.get(data_endpt, headers = {'Content-Type': 'application/json'})
        
        #tree = ET.parse(file_name)
        #root = tree.getroot()
        root = ET.fromstring(response.content)

        if (dataset == "lgg"):
            ns = {f'{dataset}': f'http://tcga.nci/bcr/xml/clinical/{dataset}/2.7',
            'shared': 'http://tcga.nci/bcr/xml/shared/2.7'}
            nextNode = root.find(f'{dataset}:patient', ns)
            tumor_stage = nextNode.find('shared:neoplasm_histologic_grade', ns).text
        else:
            ns = {f'{dataset}': f'http://tcga.nci/bcr/xml/clinical/{dataset}/2.7',
               'shared_stage': 'http://tcga.nci/bcr/xml/clinical/shared/stage/2.7'}
            nextNode = root.find(f'{dataset}:patient', ns)
            nextNode = nextNode.find('shared_stage:stage_event', ns)
            tumor_stage = nextNode.find('shared_stage:pathologic_stage', ns).text
        
        print(tumor_stage)

        return (tumor_stage)

    except:
        return DEFAULT

def get_filename(filters, DEFAULT):

    resp = rq.post(f'https://api.gdc.cancer.gov/files?size=100', json={'filters': filters})

    #print(resp.json())

    filenum = 0
    notFound = True


    while (filenum < len(resp.json()['data']['hits']) and notFound):
        file = resp.json()['data']['hits'][filenum]
        #print(file)
        if ('nationwidechildrens.org_clinical.' in file['file_name']):
            notFound = False
        filenum+=1
    
    if notFound:
        return DEFAULT

    else:
        return file['file_id']

class _FilterBuilder:

    @staticmethod
    def logical(op, args):
        ''' Logical operator '''

        _filter = { 'op': op, 'content': [o for o in args] }

        return _filter

    @staticmethod
    def inclusion(field, values):
        ''' Inclusion operator '''

        if len(values) < 1:
            raise RuntimeError(f'Invalid number of values: {len(values)}')

        _filter = { 'op': 'in', 'content': { 'field': field, 'value': values }}

        return _filter

    @staticmethod
    def equal(field, value):
        ''' Equal operator '''

        _filter = { 'op': '=', 'content': { 'field': field, 'value': value }}

        return _filter



if __name__ == '__main__':
    main()
