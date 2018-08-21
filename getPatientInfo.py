import logging as log
import pandas as pd
import requests as rq
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import sys
import time
import csv
import xlsxwriter
from openpyxl import load_workbook
import os.path

class GDC_API:

    def __init__(self, gdc_url='https://api.gdc.cancer.gov', per_page=100, logfile=None):
        """GDC API initialization
        
        Keyword Arguments:
            gdc_url {str} -- gdc api url (default: {'https://api.gdc.cancer.gov'})
            per_page {int} -- number of results per page (default: {100})
            logfile {str} -- filepath for logging (default: {None})
        """

        self.gdc_url = gdc_url
        self.per_page = per_page

        log.basicConfig(filename=logfile, level=log.INFO, format='%(asctime)s : %(levelname)8s : %(message)s (%(module)s.%(funcName)s)', datefmt='%Y-%m-%d %H:%M:%S')

    def get_filtered_case_ids(self, project, filter_field, filter_val):
        """Get SNP data formatted for MutComFocal analysis
        
        Arguments:
            projects {str} -- project name, e.g. 'TCGA-LUAD'
        
        Keyword Arguments:
            workflow {str} -- workflow for SNP data as defined in the gdc data portal (default: {'MuTect2 Variant Aggregation and Masking'})
        
        Returns:
            pandas DataFrame -- matrix of SNP data as expected by MutComFocal
        """

        log.info(f'Project {project}')

        filters = _FilterBuilder.logical('and', [
            _FilterBuilder.equal(filter_field, filter_val),
            _FilterBuilder.equal('cases.project.project_id', project)])
        
        case_ids = self._get_case_ids(filters, filter_val)
        return case_ids


    
    def _get_case_ids(self, filters, filter_val):

        # case_ids = []

        ids = []

        log.info('    Getting first page of case ids')

        resp = rq.post(f'{self.gdc_url}/cases?size={self.per_page}', json={'filters': filters})

        #f'{meta['pages']}'



        if resp.status_code == 200:
            resp = resp.json()
            meta = resp['data']['pagination']

        for h in resp['data']['hits']:
        	#case_ids += [h['case_id']]
        	samp = [h['case_id'], h['submitter_id'], filter_val]
        	ids.append(samp)

        if meta['pages'] > 1:
            for _from in range(self.per_page, meta['total'], self.per_page):
                log.info(f'    Getting page for case ids from {_from}')
                resp = rq.post(f'{self.gdc_url}/cases?size={self.per_page}&from={_from}', json={'filters': filters})
                if resp.status_code == 200:
                    resp = resp.json()

                    for h in resp['data']['hits']:
                    	#case_ids += [h['case_id']]
                    	samp = [h['case_id'], h['submitter_id'], filter_val]
                    	ids.append(samp)
                    #f'{case_ids}'

        #return case_ids
        return ids


class _FilterBuilder:

    @staticmethod
    def logical(op, args):
        ''' Logical operator '''

        _filter = { 'op': op, 'content': [o for o in args] }

        return _filter

    @staticmethod
    def inclusion(field, values):
        ''' Inclusion operator '''

        if len(values) < 1:
            raise RuntimeError(f'Invalid number of values: {len(values)}')

        _filter = { 'op': 'in', 'content': { 'field': field, 'value': values }}

        return _filter

    @staticmethod
    def equal(field, value):
        ''' Equal operator '''

        _filter = { 'op': '=', 'content': { 'field': field, 'value': value }}

        return _filter

def main():

    gdc = GDC_API()
    p = sys.argv[1]
    filter_field = sys.argv[2] #i.e. cases.demographic.gender
    filter_val = sys.argv[3:] #i.e. male

    out = []

    for val in filter_val:
        i = 0
        temp = val
        while (i < len(val)):
            if (val[i] == '_'):
                temp = temp[0:i] + ' ' + temp[i+1:]
            i+=1
        val = temp
        log.info(val)
        out = out + gdc.get_filtered_case_ids(p, filter_field, val)

    header = ['case_id', 'submitter_id', filter_field]

    filename = f'{p}-patient-data.xlsx'

    df = pd.DataFrame(data=out)

    if (os.path.isfile(filename)):
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine = 'openpyxl')
        writer.book = book
    else:
        writer = pd.ExcelWriter(filename, engine = 'xlsxwriter')

    df.to_excel(writer, header = header, sheet_name = filter_field, index = False)
    writer.save()

    # with xlsxwriter.Workbook('TGCA-BLCA.xlsx') as workbook:
    #     worksheet = workbook.add_worksheet()
    #     worksheet.set_header(header)
    #     for row, data in enumerate(out):
    #         worksheet.write_row(row, 0, data)

    # workbook.close()


    #for p in projects:
        #out = pd.DataFrame(gdc.get_filtered_case_ids(p))
        #out.to_excel('out.xlsx',index=False)
        #out.to_csv(f'{p}_ids.txt', sep='\t')
        #with open('out', 'w') as myfile:
        #    wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
            
        #    wr.writerow(out)

if __name__ == '__main__':
    main()