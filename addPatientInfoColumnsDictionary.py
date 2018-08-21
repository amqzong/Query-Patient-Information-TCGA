import xlsxwriter
from openpyxl import load_workbook
import sys
import pandas as pd
import logging as log
import os.path
import pickle

def main():

	p = sys.argv[1]
	ID_SIZE = 12
	DEFAULT = 'no data'

	try:
		uniqueIDs = pickle.load(open(f'{p}-uniqueIDs.pkl', 'rb'))
	except:
		print ("Starting to compile unique IDs")
		uniqueIDs = getUniquePatientIDs(p, ID_SIZE, DEFAULT)
		f = open(f'{p}-uniqueIDs.pkl', 'wb')
		pickle.dump(uniqueIDs, f)
		f.close()

	print(uniqueIDs)

	print ("Finished compiling unique IDs")
	matchUnique(p, uniqueIDs, ID_SIZE, DEFAULT) #add onto list

	print ("Finished creating new excel sheet")

def getUniquePatientIDs(p, ID_SIZE, DEFAULT):
	ID_COL_NUM = 1
	filename = f'{p}-raw.xlsx'

	colname = 'ID'

	uniqueIDs = {}

	wb = load_workbook(filename)
	sheet = wb.worksheets[0]

	#import ipdb; ipdb.set_trace()

	for row in range(1, sheet.max_row+1):
		newID = sheet.cell(row,ID_COL_NUM).value

		if (newID): #check if string is not null
			newID = newID[0:ID_SIZE]

			if (not(newID in uniqueIDs)):
				uniqueIDs[newID] = DEFAULT

	return uniqueIDs

def matchUnique(p, uniqueIDs, ID_SIZE, DEFAULT):
	ID_COL_NUM = 2;
	FIELD_COL_NUM = 3;
	filename = f'{p}-patient-data.xlsx'

	wb = load_workbook(filename)
	num_sheets = len(wb.worksheets)

	for i in range(1,2):
	#for i in range(0, num_sheets):
		sheet = wb.worksheets[i]
		field = wb.sheetnames[i]

		uniqueIDs = dict.fromkeys(uniqueIDs, DEFAULT)

		for row in range(1, sheet.max_row+1):
			nextData = sheet.cell(row, ID_COL_NUM).value
			if (nextData):
				nextData = nextData[0:ID_SIZE]
				if (nextData in uniqueIDs):
					uniqueIDs[nextData] = sheet.cell(row, FIELD_COL_NUM).value

		createNewColumn(p, uniqueIDs, ID_SIZE, DEFAULT, field)

def createNewColumn(p, matchIDsData, ID_SIZE, DEFAULT, field):
	ID_COL_NUM = 1;
	filename = f'{p}-added.xlsx'

	if (os.path.isfile(filename)):
		wb = load_workbook(filename)
	else:
		wb = load_workbook(f'{p}-raw.xlsx')

	sheet = wb.worksheets[0]
	maxrow = sheet.max_row
	maxcol = sheet.max_column

	colexist = False
	col = 1

	while (not (colexist) and col < maxcol):
		if (sheet.cell(1,col).value == field):
			colexist = True
		col+=1


	if (not colexist):
		sheet.cell(1,maxcol+1).value = field

		for row in range(2, maxrow+1):
			newID = sheet.cell(row,ID_COL_NUM).value

			if (newID): #check if newID is non-null
				newID = newID[0:ID_SIZE]

				if (not(newID in matchIDsData)):
					sheet.cell(row, maxcol+1).value = DEFAULT
				else:
					sheet.cell(row, maxcol+1).value = matchIDsData[newID]

		wb.save(filename)

if __name__ == '__main__':
    main()